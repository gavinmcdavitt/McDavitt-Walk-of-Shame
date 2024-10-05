from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import xlwt
from xlwt import Workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import requests
def getListOfGames():
    url = "https://www.cbssports.com/nfl/scoreboard/"  # Replace with the target URL
    response = requests.get(url)
    print(response)
    soup = BeautifulSoup(response.content, 'html.parser')
    team_links = soup.find_all('a', class_='team-name-link')
    listOfTeams=[]
    for team in team_links:
        team_name = team.get_text(strip=True)
        listOfTeams.append(team_name)
        print(team_name)
    return listOfTeams
def getPregameOdds():
    url = "https://www.cbssports.com/nfl/scoreboard/"  # Replace with the target URL
    response = requests.get(url)
    print(response)
    soup = BeautifulSoup(response.content, 'html.parser')
    over_under = soup.find_all('td', class_="in-progress-odds in-progress-odds-away")
    spreads = soup.find_all('td', class_='in-progress-odds in-progress-odds-home')
    listofBetting =[]
    for a , b in zip(over_under, spreads):
        AwayOverUnder = a.get_text(strip=True)
        homeSpread = b.get_text(strip=True)
       # print(AwayOverUnder, " ", homeSpread)
        bets = {
            'over_under': AwayOverUnder,
            'spread': homeSpread
        }
        listofBetting.append(bets)
    return listofBetting
    # for a, b in over_under and spreads:
    #     AwayOverUnder = a.get_text(strip=True)
    #     homeSpread = b.get_text(strip=True)
    #     print('over-under: ', AwayOverUnder, 'spread ', homeSpread)

    #class="in-progress-odds in-progress-odds-away"

def getGames(teams, odds):
    listOfCompetitions = []

    # Iterate over the teams in pairs (home and away)
    for i in range(0, len(teams), 2):
        # Calculate the corresponding index in the odds list
        odds_index = i // 2

        # Check if there's a valid odds entry for this game
        if odds_index < len(odds):
            # Create a competition entry for this pair of teams and odds
            competition = {
                'home': teams[i],  # Home team
                'away': teams[i + 1],  # Away team
                'spread': odds[odds_index]['spread'],  # Corresponding spread
                'over_under': odds[odds_index]['over_under']  # Corresponding over/under
            }
            listOfCompetitions.append(competition)
        else:
            print(f"Warning: Not enough odds for game between {teams[i]} and {teams[i + 1]}")
    print('amount of games: ', odds_index)
    return listOfCompetitions

def getNumOfGames(run):
    return len(run)

def printToExcel(data, weekNum):

    # Create a new Workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Optionally, you can set a title for the sheet
    ws.title = "My Items"

    # Write headers (optional)
    ws['A1'] = 'Away'  # Header for Column A
    ws['B1'] = 'Home'  # Header for Column B
    ws['C1'] = 'spread'
    ws['D1'] = 'Over Under'
    ws['E1'] ='Pops'
    ws['F1'] ='Anna'
    ws['G1'] = 'Gavin'
    ws['H1'] ='Winner'

    # Write data to column B starting from the second row
    for index, item in enumerate(data, start=1):  # Start from row 2
        ws[f'A{index+1}'] = item['home']  # Write away team
        ws[f'B{index+1}'] = item['away']  # Write home team
        ws[f'C{index+1}'] = item['spread']  # Write Spread
        ws[f'D{index+1}'] = item['over_under']  # Write Over Under

    # Save the workbook to a file
    excel_file_path = F"week{weekNum}.xlsx"
    wb.save(excel_file_path)

    print(f"Data written to {excel_file_path}")
    return excel_file_path

def ColorizePicks(fileName, num):
    #if there are the same picks for everyone highlight them all to be green.
    wb = load_workbook(fileName)
    ws = wb.active
    #you need to grab e2 - g2.
    #create THREE variables that will be fStrings to grab the cells to edit.
    for i in range(2, 2+num):
        ColE=f"E{i}"
        ColF = f"F{i}"
        ColG =f"G{i}"
        if ws[ColE].value == ws [ColF].value == ws[ColG].value:
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            ws[ColE].fill = green_fill
            ws[ColF].fill = green_fill
            ws[ColG].fill = green_fill
        red_Fill = PatternFill(start_color="FFA500", end_color="FF0000", fill_type="solid")
        if ws[ColE].value != ws[ColF].value:

            ws[ColE].fill = red_Fill
            ws[ColF].fill = red_Fill
            #ws[ColG].fill = red_Fill
        if ws[ColE].value != ws[ColG].value:

            ws[ColE].fill = red_Fill
            ws[ColG].fill = red_Fill
        if ws[ColF].value != ws[ColG].value:

            ws[ColF].fill = red_Fill
            ws[ColG].fill = red_Fill
        wb.save(fileName)
        return fileName
def setUp():
    #only do before the games start. After thursday.
    weekNumber =5
    teams = getListOfGames()
    odds = getPregameOdds()
    run = getGames(teams, odds)
    num = getNumOfGames(run)
    fileName = printToExcel(run, weekNumber)

def getWinners(weekNumber):
    url = f"https://www.cbssports.com/nfl/scoreboard/all/2024/regular/{weekNumber}/"  # Replace with the target URL
    response = requests.get(url)
    print(response)
    soup = BeautifulSoup(response.content, 'html.parser')
    teams = soup.find_all('a', class_="team-name-link")
    competition={}
    listOfTeams = []
    itr = 0
    isHome= False
    for team in teams:
        team_name = team.get_text(strip=True)
        listOfTeams.append(team_name)

    scores = soup.find_all('td', class_="total")
    listOfScores =[]
    for points in scores:
        sco = points.get_text(strip=True)
        print(sco)
        listOfScores.append(sco)

    for i in range(0, len(listOfTeams), 2):
        index = i//2
        listofGames = []
        winner =""
        loser = ""
        winnerPoints =0
        loserPoints =0
        if index < len(listOfTeams):
            if int(listOfScores[i]) < int(listOfScores[i+1]):
                winnerPoints = listOfScores[i+1]
                loserPoints = listOfScores[i]
                winner = listOfTeams[i+1]
                loser = listOfTeams[i]
            elif int(listOfScores[i+1]) < int(listOfScores[i]):
                winnerPoints = listOfScores[i]
                loserPoints = listOfScores[i+1]
                winner = listOfTeams[i]
                loser = listOfTeams[i+1]

            competition={
                'Winner':winner,
                'Loser':loser,
                'Loser-points':loserPoints,
                'Winner-points':winnerPoints
            }
            print(competition)
            listofGames.append(competition)




#setUp()
#fileName =ColorizePicks(fileName='week4.xlsx',num =15)
#getWinners(4)
getWinners(4)

